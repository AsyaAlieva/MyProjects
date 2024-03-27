import random
import math
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

iteration = 1

#Функция для моделирования времени работы
def WorkTime(machine):
    if machine == "Самосвал":
        return -4*math.log(1-random.random())
    else:
        return -6*math.log(1-random.random())

B = "Бульдозер"
C = "Самосвал"
# Функция для моделирования времени ремонта
def RepairTime(level, machine):
    B = "Бульдозер"
    C = "Самосвал"
    if level == 3:
        if machine == C:
            return -2*math.log(1-random.random())
        else:
            return None
    if level == 6:
        if machine == C:
            return -1*math.log(1-random.random())
        elif machine == B:
            return -2*math.log(1-random.random())
    elif level == 36:
        if machine == C:
            return -0.25*math.log(1-random.random())
        elif machine == B:
            return -1.5*math.log(1-random.random())

# Моделирование рабочего дня
# Функция для моделирования 1000 дней
def simulate_days(rank): #с помощью for цикл по 1000 итерациям и с помощью while цикл где граниченое значение 16 (ч)
    work_sum_C = 0
    repair_sum_C = 0
    waiting_sum_C = 0

    work_sum_B = 0
    repair_sum_B = 0
    waiting_sum_B = 0

    for i in range(1, 1001):

        B_state = {'Работает': None,
                    'Простой': None,
                    'На ремонте': None}
        C_state = {'Работает': None,
                    'Простой': None,
                    'На ремонте': None}

        Bgeneral = [] #здесь будут хранится все состояния по одному дню для бульдозера
        Cgeneral = [] #здесь будут хранится все состояния по одному дню для самосвала

        timeB = 0  # общее время (работы + простоя + ремонта) бульдозера
        timeC = 0  # общее время (работы + простоя + ремонта) самосвала

        C_Status_wait_rep = False #статус ожидания ремонта самосвалом
        B_Status_wait_rep = False #статус ожидания ремонта бульдозером


        while (timeB < 16) or (timeC < 16):
            # Время работы Самсовала
            if C_Status_wait_rep == False: # если Самосвал не ждет ремонта
                C_work = WorkTime(C) #высчитываем значение работы Самосвала по функции (к-рая считается по эксп. распр-ю)
                if timeC < 16:
                    if timeC + C_work > 16: #если последнее получаемое значение больше чем границы (т.е чем 16)
                        C_work = 16 - timeC # мы отнимаем от 16 количество, которое уже занятое время, чтобы получить оставшийся последний кусочек времени
                    C_state['Работает'] = (timeC, timeC + C_work) #засовываем промежуток времени когда Самосвал работает - в словарь
                    work_sum_C += C_work
                    Cgeneral.append({'Работает': C_state['Работает']})
                    timeC += C_work # добавляем к общему времени уже пройденное (занятое) время



            # Время работы Бульдозера
            if B_Status_wait_rep == False: # аналогично, только работаем с Бульдозером
                B_work = WorkTime(B)
                if timeC < 16:
                    if timeB + B_work > 16:
                        B_work = 16 - timeB
                        work_sum_B += B_work
                    B_state['Работает'] = (timeB, timeB + B_work)
                    work_sum_B += B_work
                    Bgeneral.append({'Работает': B_state['Работает']})
                    timeB += B_work


            # Время когда нужно отремонтировать машину (т.е Самосвал или Бульдозер)
            C_repair = RepairTime(rank, C) # время ремонта Самосвала
            B_repair = RepairTime(rank, B) # время ремонта Бульдозера

            if timeC <= timeB: # если общее время Самосвала на данный момент меньше времени Бульдозера, считаем ремонт для Самосвала
                C_state['На ремонте'] = (timeC, min(timeC+C_repair, 16)) #высчитываем координаты на временной шкале
                repair_sum_C += C_repair
                Cgeneral.append({'На ремонте': C_state['На ремонте']})
                timeC = min(timeC+C_repair, 16) #конечная точка на временной шкале на данный момент времени является текущим значением времени
                C_Status_wait_rep = False #мы посчитали и учли время ремонта, => Самосвал теперь отремонтирован

                #начинаем разбираться, есть ли простой (у Бульдозерра)
                if timeB < timeC: #когда у нас Самосвал ремонтируется, а Бульдозер уже поработал и хочет отремонтироваться
                    wait_value = timeC - timeB # получаем значение простоя для Бульдозера
                    B_state['Простой'] = (timeB, timeB+wait_value) #записываем координаты на временной шкале
                    waiting_sum_B += wait_value
                    B_state['На ремонте'] = (timeB+wait_value, min(timeB+wait_value+B_repair, 16)) #ремонт Бульдозера начинается
                    #когда отремонтировался Самосвал
                    Bgeneral.append({'Простой': B_state['Простой']})
                    Bgeneral.append({'На ремонте': B_state['На ремонте']})
                    repair_sum_B += B_repair
                    timeB = min(timeB+wait_value+B_repair, 16) #получаем значение текущего времени для Бульдозера
                else:
                    B_Status_wait_rep = True # не разрешит присвоить бульдозеру значение ремонта на врем. шкале, т.к. оно может
                    #перекрыться на шкале возможно предстоящим ремонтом самосвалом

            else: #если обратная ситуация, считаем ремонт для Бульдозера
                B_state['На ремонте'] = (timeB, min(timeB + B_repair, 16))
                repair_sum_B += B_repair
                Bgeneral.append({'На ремонте': B_state['На ремонте']})
                timeB = min(timeB+B_repair, 16)
                B_Status_wait_rep = False

                # начинаем разбираться, есть ли простой (у Самосвала)
                if timeC < timeB:
                    wait_value = timeB - timeC
                    C_state['Простой'] = (timeC, timeC + wait_value)  # записываем координаты на временной шкале
                    C_state['На ремонте'] = (timeC + wait_value, min(timeC + wait_value + C_repair, 16))
                    waiting_sum_C += wait_value
                    repair_sum_C += C_repair
                    Cgeneral.append({'Простой': C_state['Простой']})
                    Cgeneral.append({'На ремонте': C_state['На ремонте']})
                    timeC = min(timeC+wait_value+C_repair, 16)
                else:
                    C_Status_wait_rep = True

        if i == iteration: # Выбираем по какому дню хотим получить график

            listC_state = [] #состояния Самосвала
            listC_values_start= [] #Начальные координаты самосвала по всем состояним
            listC_values_end = [] # Конечные координаты самосвала по всем состояниям
            listB_state = []
            listB_values_start = []
            listB_values_end = []

            for d1 in Cgeneral:
                for key1, value1 in d1.items():
                    listC_state.append(key1)
                    listC_values_start.append(value1[0])
                    listC_values_end.append(value1[1])

            for d2 in Bgeneral:
                for key2, value2 in d2.items():
                    listB_state.append(key2)
                    listB_values_start.append(value2[0])
                    listB_values_end.append(value2[1])

            # подключаем эксель файл, очищаем его от старых данных и выгружаем туда необходимые нам данные
            file = 'chart.xlsx'
            wb = load_workbook(file)
            wb.remove(wb['data_C'])
            wb.remove(wb['data_B'])
            ws1 = wb.create_sheet(title='data_C')
            ws2 = wb.create_sheet(title='data_B')
            wb.save(filename='chart.xlsx')

            # Заполняем файл данными для Самосвала
            for row in range(1, len(listC_state)+1):
                value1 = listC_state[row-1]
                cell = ws1.cell(row=row, column=1)
                cell.value = value1
            for row in range(1, len(listC_values_start)+1):
                value2 = listC_values_start[row-1]
                cell = ws1.cell(row=row, column=2)
                cell.value = value2
            for row in range(1, len(listC_values_end)+1):
                value3 = listC_values_end[row-1]
                cell = ws1.cell(row=row, column=3)
                cell.value = value3
            # Формируем для график Самосвала
            values = Reference(ws1, min_col=1, min_row=1, max_col=3, max_row=len(listC_state))
            chart = BarChart()
            chart.add_data(values)
            ws1.add_chart(chart, "F4")


            #Заполняем файл данными для Бульдозера
            for row in range(1, len(listB_state)+1):
                value = listB_state[row-1]
                cell = ws2.cell(row=row, column=1)
                cell.value = value
            for row in range(1, len(listB_values_start)+1):
                value = listB_values_start[row-1]
                cell = ws2.cell(row=row, column=2)
                cell.value = value
            for row in range(1, len(listB_values_end)+1):
                value = listB_values_end[row-1]
                cell = ws2.cell(row=row, column=3)
                cell.value = value
            #Формируем для график Бульдозера
            # from openpyxl.chart import BarChart, StockChart, Reference, Series
            values = Reference(ws2, min_col=1, min_row=1, max_col=3, max_row=len(listB_state))
            chart = BarChart()
            chart.add_data(values)
            ws2.add_chart(chart, "F4")

            #Сохраняем и закрываем файл
            wb.save(file)
            wb.close()

        # print("день", i)
        # print(f"САМОСВАЛ - {Cgeneral}")
        # print(f"БУЛЬДОЗЕР - {Bgeneral}")

    if rank == 6:
        expenses = (waiting_sum_C * 1500 + waiting_sum_B * 1300) + ((repair_sum_C + repair_sum_B) * (900+100))
        income = work_sum_B * 1500 + work_sum_C * 1300
        profit = income - expenses
    elif rank == 36:
        expenses = (waiting_sum_C * 1500 + waiting_sum_B * 1300) + ((repair_sum_C + repair_sum_B) * (900+600+100))
        income = work_sum_B * 1500 + work_sum_C * 1300
        profit = income - expenses

    return profit


p1 = simulate_days(6)
p2 = simulate_days(36)

print(f"Прибыль когда работает слесарь 6-го разряда: {p1} рублей")
print(f"Прибыль когда работает бригада из слесарей 3-го и 6-го разрядов: {p2} рублей")
if p2 > p1:
    print(f"Не выгодно уволнять слесаря 3-го разряда, т.к иначе прибыль будет меньше на: {round(p2-p1)} рублей")
else:
    print(f"Выгодно уволить слесаря 3-го разряда, т.к мы не дополучим: {round(p1-p2)} рублей")




